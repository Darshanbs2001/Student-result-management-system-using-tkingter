from tkinter import *
from create_db import *
from openpyxl import Workbook,load_workbook
from tkinter import ttk,messagebox
wb=load_workbook("hello1.xlsx")
ws=wb.active
class Result:
    def __init__(self,root):
        self.root = root
        self.root.title("Student Result Mangament System")
        self.root.geometry("1200x480+70+170")
        self.root.config(bg="white")
        self.root.focus_force()
        # ====title===
        title = Label(self.root, text="Add Student Result", font=("goudy old style", 20, "bold"), bg="gold",
                      fg="black").place(x=0, y=15, width=1200, height=35)
      #=====variable================
        self.rollno=StringVar()
        self.semister=StringVar()
        self.var_tosubject1=StringVar()
        self.var_tmaxsubject1=StringVar()
        self.var_pmaxsubject1=StringVar()
        self.var_posubject1=StringVar()
        self.var_tosubject2=StringVar()
        self.var_tmaxsubject2=StringVar()
        self.var_pmaxsubject2=StringVar()
        self.var_posubject2=StringVar()
        self.var_tmaxsubject3=StringVar()
        self.var_tosubject3=StringVar()
        self.var_pmaxsubject3=StringVar()
        self.var_posubject3=StringVar()



        self.lbl_rollno=Label(self.root,text="Enter the Student Roll No:",font=("goudy old style",15,"bold"),fg="black",bg="white").place(x=10,y=60)
        self.txt_rollno=Entry(self.root,text=self.rollno,font=("goudy old style",15,"bold"),fg="black",bg="lightyellow").place(x=240,y=60,width=200)
      #  lbl_semister = Label(self.root, text="Enter the Student Roll Number :", font=("goudy old style", 20), fg="black", lbl_rollno=Label(self.root,text="Enter the Student Roll No:",font=("goudy old style",15,"bold"),fg="black",bg="white").place(x=10,y=60)
        self.lbl_semister = Label(self.root, text="Select Semister", font=("goudy old style", 15, "bold"),
                           fg="black", bg="white").place(x=460,y=60)
        self.txt_semister = ttk.Combobox(self.root, textvariable=self.semister, values=["Select","I-Semister","II-Semister","III-Semister","IV-Semister","V-Semister","VI-Semister"],
                                       font=("goudy old style ", 15, "bold"), justify="center", state="readonly").place(x=600,y=60,width=200)
        self.btn_search=Button(self.root,text="Search",font=("goudy old style",15,"bold"),fg="white",bg="#03a9f4",command=self.search).place(x=840,y=60,width=200,height=35)


    #  txt_semister=Entry(self.root,text=self.search,font=("goudy old style",15,"bold"),fg="black",bg="lightyellow").place(x=230,y=60,width=200)
    #=============functions=============
    def search(self):
        try:
            if self.rollno.get()=="":
                messagebox.showerror("Error","Please enter the student register number",parent=self.root)
            else:
                mycursor.execute("select regno,sem from result where regno=%s and sem=%s",(self.rollno.get(),self.semister.get()))
                rows=mycursor.fetchall()
                print(rows)
                mycursor.execute("select roll from student where roll=%s",(self.rollno.get(),))
                srows=mycursor.fetchall()
                if srows!=[]:
                    if rows==[]:
                        self.clear()
                        mycursor.execute("select course from student where roll=%s",(self.rollno.get(),))
                        course=mycursor.fetchall()
                        print(course)
                        if course==None:
                            messagebox.showerror("Error","Student record is not in the database",parent=self.root)
                        else:
                            for c in course:
                                print(c)
                                mycursor.execute("select s1,s2,s3 from course where name=%s",(c[0],))
                                crows=mycursor.fetchall()
                                for crow in crows:
                                    self.subject1name=crow[0]
                                    self.subject2name=crow[1]
                                    self.subject3name=crow[2]
                                    self.frame_subject1=LabelFrame(self.root,text=crow[0],font=("goudy old style ",15,"bold"),bg="white",fg="black").place(x=10,y=100,width=800,height=100)
                                    btn_save=Button(self.root,text="Save",font=("goudy old style",15,"bold"),bg="#2196f3",fg="white",cursor="hand2",command=self.save).place(x=850,y=360,width=200,height=35)
                                    btn_clear=Button(self.root,text="Clear",font=("goudy old style",15,"bold"),bg="#2196f3",fg="white",cursor="hand2",command=self.clear).place(x=850,y=320,width=200,height=35)
                                    self.frame_subject2=LabelFrame(self.root,text=crow[1],font=("goudy old style ",15,"bold"),bg="white",fg="black").place(x=10,y=200,width=800,height=100)
                                    self.frame_subject3=LabelFrame(self.root,text=crow[2],font=("goudy old style ",15,"bold"),bg="white",fg="black").place(x=10,y=300,width=800,height=100)
                                    self.lbl_tmax=Label(self.root,text="Max.Theory marks",font=("goudy old style",15,"bold"),bg="white",fg="black").place(x=20,y=120)
                                    self.txt_tmax=Entry(self.root,text=self.var_tmaxsubject1,font=("goudy old style",15,"bold"),bg="white",fg="black").place(x=20,y=160,width=150,height=35)
                                    self.lbl_tobtianed = Label(self.root, text="Obtained marks", font=("goudy old style", 15, "bold"),
                                                 bg="white", fg="black").place(x=220, y=120)
                                    self.txt_tobtained = Entry(self.root, text=self.var_tosubject1,
                                                     font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=220,
                                                                                                                         y=160,
                                                                                                                         width=150,
                                                                                                                         height=35)
                                    self.lbl_pmax = Label(self.root, text="Max.Practical Marks", font=("goudy old style", 15, "bold"),
                                                 bg="white", fg="black").place(x=420, y=120)
                                    self.txt_pmax = Entry(self.root, text=self.var_pmaxsubject1,
                                                     font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=420,
                                                                                                                         y=160,
                                                                                                                         width=150,
                                                                                                                         height=35)
                                    self.lbl_pobtained = Label(self.root, text="Otained marks", font=("goudy old style", 15, "bold"),
                                                 bg="white", fg="black").place(x=620, y=120)
                                    self.txt_potained = Entry(self.root, text=self.var_posubject1,
                                                     font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=620,
                                                                                                                         y=160,
                                                                                                                         width=150,
                                                                                                                         height=35)
                                    self.lbl_tmax=Label(self.root,text="Max.Theory marks",font=("goudy old style",15,"bold"),bg="white",fg="black").place(x=20,y=220)
                                    self.txt_tmax=Entry(self.root,text=self.var_tmaxsubject2,font=("goudy old style",15,"bold"),bg="white",fg="black").place(x=20,y=260,width=150,height=35)
                                    self.lbl_tobtianed = Label(self.root, text="Obtained marks", font=("goudy old style", 15, "bold"),
                                                 bg="white", fg="black").place(x=220, y=220)
                                    self.txt_tobtained = Entry(self.root, text=self.var_tosubject2,
                                                     font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=220,
                                                                                                                         y=260,
                                                                                                                         width=150,
                                                                                                                         height=35)
                                    self.lbl_pmax = Label(self.root, text="Max.Practical Marks", font=("goudy old style", 15, "bold"),
                                                 bg="white", fg="black").place(x=420, y=220)
                                    self.txt_pmax = Entry(self.root, text=self.var_pmaxsubject2,
                                                     font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=420,
                                                                                                                         y=260,
                                                                                                                         width=150,
                                                                                                                         height=35)
                                    self.lbl_pobtained = Label(self.root, text="Otained marks", font=("goudy old style", 15, "bold"),
                                                 bg="white", fg="black").place(x=620, y=220)
                                    self.txt_potained = Entry(self.root, text=self.var_posubject2,
                                                     font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=620,
                                                                                                                         y=260,
                                                                                                                         width=150,
                                                                                                                         height=35)
                                    self.lbl_tmax = Label(self.root, text="Max.Theory marks",
                                                     font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=20,
                                                                                                                         y=320)
                                    self.txt_tmax = Entry(self.root, text=self.var_tmaxsubject3,
                                                     font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=20,
                                                                                                                         y=360,
                                                                                                                         width=150,
                                                                                                                         height=35)
                                    self.lbl_tobtianed = Label(self.root, text="Obtained marks",
                                                          font=("goudy old style", 15, "bold"),
                                                          bg="white", fg="black").place(x=220, y=320)
                                    self.txt_tobtained = Entry(self.root, text=self.var_tosubject3,
                                                          font=("goudy old style", 15, "bold"), bg="white", fg="black").place(
                                        x=220,
                                        y=360,
                                        width=150,
                                        height=35)
                                    self.lbl_pmax = Label(self.root, text="Max.Practical Marks",
                                                     font=("goudy old style", 15, "bold"),
                                                     bg="white", fg="black").place(x=420, y=320)
                                    self.txt_pmax = Entry(self.root, text=self.var_pmaxsubject3,
                                                     font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=420,
                                                                                                                         y=360,
                                                                                                                         width=150,
                                                                                                                         height=35)
                                    self.lbl_pobtained = Label(self.root, text="Otained marks",
                                                          font=("goudy old style", 15, "bold"),
                                                          bg="white", fg="black").place(x=620, y=320)
                                    self.txt_potained = Entry(self.root, text=self.var_posubject3,
                                                         font=("goudy old style", 15, "bold"), bg="white", fg="black").place(
                                        x=620,
                                        y=360,
                                        width=150,
                                        height=35)
                    else:
                        #messagebox.showinfo("Attention","the result already exists if you if wish you can update it",parent=self.root
                        mycursor.execute("select * from result where regno=%s and sem=%s",(self.rollno.get(),self.semister.get()))
                        rows=mycursor.fetchall()
                        print(rows)
                        for row in rows:
                            self.rollno.set(row[0])
                            self.semister.set(row[1])
                            self.var_tmaxsubject1.set(row[2])
                            self.var_tmaxsubject2.set(row[3])
                            self.var_tmaxsubject3.set(row[4])
                            self.var_tosubject1.set(row[5])
                            self.var_tosubject2.set(row[6])
                            self.var_tosubject3.set(row[7])
                            self.var_pmaxsubject1.set(row[8])
                            self.var_pmaxsubject2.set(row[9])
                            self.var_pmaxsubject3.set(row[10])
                            self.var_posubject1.set(row[11])
                            self.var_posubject2.set(row[12])
                            self.var_posubject3.set(row[13])
                            self.content=row[14]
                        mycursor.execute("select course from student where roll=%s", (self.rollno.get(),))
                        course = mycursor.fetchall()
                        print(course)
                        for c in course:
                            print(c)
                            mycursor.execute("select s1,s2,s3 from course where name=%s", (c[0],))
                            crows = mycursor.fetchall()
                            for crow in crows:
                                self.subject1name=crow[0]
                                self.subject2name=crow[1]
                                self.subject3name=crow[2]
                                self.frame_subject1 = LabelFrame(self.root, text=crow[0], font=("goudy old style ", 15, "bold"),
                                                            bg="white", fg="black").place(x=10, y=100, width=800, height=100)
                                self.btn_save = Button(self.root, text="Update", font=("goudy old style", 15, "bold"), bg="#2196f3",
                                                  fg="white", cursor="hand2",command=self.update).place(x=850, y=360, width=200,
                                                                                                       height=35)
                                self.btn_clear = Button(self.root, text="Clear", font=("goudy old style", 15, "bold"),
                                                  bg="#2196f3",
                                                  fg="white", cursor="hand2", command=self.clear).place(x=850, y=320,
                                                                                                         width=200,
                                                                                                         height=35)
                                self.frame_subject2 = LabelFrame(self.root, text=crow[1], font=("goudy old style ", 15, "bold"),
                                                            bg="white", fg="black").place(x=10, y=200, width=800, height=100)
                                self.frame_subject3 = LabelFrame(self.root, text=crow[2], font=("goudy old style ", 15, "bold"),
                                                            bg="white", fg="black").place(x=10, y=300, width=800, height=100)
                                self.lbl_tmax = Label(self.root, text="Max.Theory marks", font=("goudy old style", 15, "bold"),
                                                 bg="white", fg="black").place(x=20, y=120)
                                self.txt_tmax = Entry(self.root, text=self.var_tmaxsubject1, font=("goudy old style", 15, "bold"),
                                                 bg="white", fg="black").place(x=20, y=160, width=150, height=35)
                                self.lbl_tobtianed = Label(self.root, text="Obtained marks", font=("goudy old style", 15, "bold"),
                                                      bg="white", fg="black").place(x=220, y=120)
                                self.txt_tobtained = Entry(self.root, text=self.var_tosubject1,
                                                      font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=220,
                                                                                                                          y=160,
                                                                                                                          width=150,
                                                                                                                          height=35)
                                self.lbl_pmax = Label(self.root, text="Max.Practical Marks", font=("goudy old style", 15, "bold"),
                                                 bg="white", fg="black").place(x=420, y=120)
                                self.txt_pmax = Entry(self.root, text=self.var_pmaxsubject1,
                                                 font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=420,
                                                                                                                     y=160,
                                                                                                                     width=150,
                                                                                                                     height=35)
                                self.lbl_pobtained = Label(self.root, text="Otained marks", font=("goudy old style", 15, "bold"),
                                                      bg="white", fg="black").place(x=620, y=120)
                                self.txt_potained = Entry(self.root, text=self.var_posubject1,
                                                     font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=620,
                                                                                                                         y=160,
                                                                                                                         width=150,
                                                                                                                         height=35)
                                self.lbl_tmax = Label(self.root, text="Max.Theory marks", font=("goudy old style", 15, "bold"),
                                                 bg="white", fg="black").place(x=20, y=220)
                                self.txt_tmax = Entry(self.root, text=self.var_tmaxsubject2, font=("goudy old style", 15, "bold"),
                                                 bg="white", fg="black").place(x=20, y=260, width=150, height=35)
                                self.lbl_tobtianed = Label(self.root, text="Obtained marks", font=("goudy old style", 15, "bold"),
                                                      bg="white", fg="black").place(x=220, y=220)
                                self.txt_tobtained = Entry(self.root, text=self.var_tosubject2,
                                                      font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=220,
                                                                                                                          y=260,
                                                                                                                          width=150,
                                                                                                                          height=35)
                                self.lbl_pmax = Label(self.root, text="Max.Practical Marks", font=("goudy old style", 15, "bold"),
                                                 bg="white", fg="black").place(x=420, y=220)
                                self.txt_pmax = Entry(self.root, text=self.var_pmaxsubject2,
                                                 font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=420,
                                                                                                                     y=260,
                                                                                                                     width=150,
                                                                                                                     height=35)
                                self.lbl_pobtained = Label(self.root, text="Otained marks", font=("goudy old style", 15, "bold"),
                                                      bg="white", fg="black").place(x=620, y=220)
                                self.txt_potained = Entry(self.root, text=self.var_posubject2,
                                                     font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=620,
                                                                                                                         y=260,
                                                                                                                         width=150,
                                                                                                                         height=35)
                                self.lbl_tmax = Label(self.root, text="Max.Theory marks",
                                                 font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=20,
                                                                                                                     y=320)
                                self.txt_tmax = Entry(self.root, text=self.var_tmaxsubject3,
                                                 font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=20,
                                                                                                                     y=360,
                                                                                                                     width=150,
                                                                                                                     height=35)
                                self.lbl_tobtianed = Label(self.root, text="Obtained marks",
                                                      font=("goudy old style", 15, "bold"),
                                                      bg="white", fg="black").place(x=220, y=320)
                                self.txt_tobtained = Entry(self.root, text=self.var_tosubject3,
                                                      font=("goudy old style", 15, "bold"), bg="white", fg="black").place(
                                    x=220,
                                    y=360,
                                    width=150,
                                    height=35)
                                self.lbl_pmax = Label(self.root, text="Max.Practical Marks",
                                                 font=("goudy old style", 15, "bold"),
                                                 bg="white", fg="black").place(x=420, y=320)
                                self.txt_pmax = Entry(self.root, text=self.var_pmaxsubject3,
                                                 font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=420,
                                                                                                                     y=360,
                                                                                                                     width=150,
                                                                                                                     height=35)
                                self.lbl_pobtained = Label(self.root, text="Otained marks",
                                                      font=("goudy old style", 15, "bold"),
                                                      bg="white", fg="black").place(x=620, y=320)
                                self.txt_potained = Entry(self.root, text=self.var_posubject3,                            font=("goudy old style", 15, "bold"), bg="white", fg="black").place(x=620,y=360,width=150,height=35)

                else:
                                messagebox.showerror("Error","Student is not registered",parent=self.root)


        except EXCEPTION as ex:
            messagebox.showerror("Error", f"Error due to {str(ex)}")

    def add_to_workbook(self):
        ws['A3']=self.subject1name
        ws['B3'] =int(self.var_tmaxsubject1.get())
        ws['C3'] =int(self.var_tosubject1.get())
        ws['D3'] =int(self.var_pmaxsubject1.get())
        ws['E3'] =int(self.var_posubject1.get())
        ws['F3']=int(self.var_tosubject1.get())+int(self.var_posubject1.get())
        ws['A4']=self.subject2name
        ws['B4'] =int(self.var_tmaxsubject2.get())
        ws['C4'] =int(self.var_tosubject2.get())
        ws['D4'] =int(self.var_pmaxsubject2.get())
        ws['E4'] =int(self.var_posubject2.get())
        ws['F4'] = int(self.var_tosubject2.get()) + int(self.var_posubject2.get())
        ws['A5']=self.subject3name
        ws['B5'] =int(self.var_tmaxsubject3.get())
        ws['C5'] =int(self.var_tosubject3.get())
        ws['D5'] =int(self.var_pmaxsubject3.get())
        ws['E5'] =int(self.var_posubject3.get())
        ws['F5'] = int(self.var_tosubject3.get()) + int(self.var_posubject3.get())

        n = 0
        maxtotal=int(self.var_tmaxsubject1.get())+int(self.var_pmaxsubject1.get())+int(self.var_tmaxsubject2.get())+int(self.var_pmaxsubject2.get())+int(self.var_tmaxsubject3.get())+int(self.var_pmaxsubject3.get())
        sum=int(self.var_tosubject1.get())+int(self.var_posubject1.get())+int(self.var_tosubject2.get())+int(self.var_posubject2.get())+int(self.var_tosubject3.get())+int(self.var_posubject3.get())
        per=float(sum/maxtotal)*100
        ws['F6']=sum
        ws['F7']=per
        if int(self.var_tosubject1.get()) >=( 0.21 * int(self.var_tmaxsubject1.get())) and int(self.var_posubject1.get()) >= (0.5 * int(self.var_pmaxsubject1.get())):
            ws['G3'] = "PASS"
            n = n + 1;
        else:
            ws['G3'] = "FAIL"

        if int(self.var_tosubject2.get()) >= (0.21 * int(self.var_tmaxsubject2.get())) and int(self.var_posubject2.get()) >= (0.5 * int(self.var_pmaxsubject2.get())):
            ws['G4'] = "PASS"
            n = n + 1;
        else:
            ws['G4'] = "FAIL"

        if int(self.var_tosubject3.get()) >= (0.21 * int(self.var_tmaxsubject3.get())) and int(self.var_posubject3.get()) >= (0.5 * int(
                self.var_pmaxsubject3.get())):
            ws['G5'] = "PASS"
            n = n + 1;
        else:
            ws['G5'] = "FAIL"
        if n == 3:
            ws['F8'] = "PASS"
        else:
            ws['F8'] = "FAIL"
        wb.save("result2.xlsx")
        fo = open("result2.xlsx", 'rb')
        content = fo.read()
        return(content)
    def save(self):
        try:

            mycursor.execute("select regno,sem from result where regno=%s and sem=%s",(self.rollno.get(),self.semister.get(),))
            rows=mycursor.fetchall()
            mycursor.execute("select course from student where roll=%s",(self.rollno.get(),))
            crows=mycursor.fetchall()
            for c in crows:
                course=c[0]
            if rows==[]:
                content=self.add_to_workbook()
                mycursor.execute("insert into result(regno,sem,tmax1,tmax2,tmax3,tm1,tm2,m3,pmax1,pmax2,pmax3,pm1,pm2,pm3,result_file,course) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",(self.rollno.get(),self.semister.get(),self.var_tmaxsubject1.get(),self.var_tmaxsubject2.get(),self.var_tmaxsubject3.get(),self.var_tosubject1.get(),self.var_tosubject2.get(),self.var_tosubject3.get(),self.var_pmaxsubject1.get(),self.var_pmaxsubject2.get(),self.var_pmaxsubject3.get(),self.var_posubject1.get(),self.var_posubject2.get(),self.var_posubject3.get(),content,course))
                mydb.commit()

                messagebox.showinfo("Success","The result has successfully been inserted",parent=self.root)
            else:
                messagebox.showerror("Error","Result is already exists",parent=self.root)
        except EXCEPTION as ex:
            messagebox.showerror("Error",f"Error due to {str(ex)}")
    def update(self):
        try:
            content=self.add_to_workbook()
            mycursor.execute("select course from student where roll=%s",(self.rollno.get(),))
            crows=mycursor.fetchall()
            for c in crows:
                course=c[0]
            mycursor.execute("update result set tmax1=%s,tmax2=%s,tmax3=%s,tm1=%s,tm2=%s,m3=%s,pmax1=%s,pmax2=%s,pmax3=%s,pm1=%s,pm2=%s,pm3=%s,result_file=%s,course=%s where regno=%s and sem=%s",
                             (
                                 self.var_tmaxsubject1.get(),
                                 self.var_tmaxsubject2.get(),
                                 self.var_tmaxsubject3.get(),
                                 self.var_tosubject1.get(),
                                 self.var_tosubject2.get(),
                                 self.var_tosubject3.get(),
                                 self.var_pmaxsubject1.get(),
                                 self.var_pmaxsubject2.get(),
                                 self.var_pmaxsubject3.get(),
                                 self.var_posubject1.get(),
                                 self.var_posubject2.get(),
                                 self.var_posubject3.get(),
                                 content,
                                 course,
                                 self.rollno.get(),
                                 self.semister.get(),

                             )
                             )
            mydb.commit()
            messagebox.showinfo("Success","The result has been updated",parent=self.root)
        except EXCEPTION as ex:
            messagebox.showerror("Error",f"Error due to {str(ex)}")
    def clear(self):
        self.var_tmaxsubject1.set("")
        self.var_tmaxsubject2.set("")
        self.var_tmaxsubject3.set("")
        self.var_tosubject1.set("")
        self.var_tosubject2.set("")
        self.var_tosubject3.set("")
        self.var_pmaxsubject1.set("")
        self.var_pmaxsubject2.set("")
        self.var_pmaxsubject3.set("")
        self.var_posubject1.set("")
        self.var_posubject2.set("")
        self.var_posubject3.set("")

if __name__=="__main__":
    root=Tk()
    obj=Result(root)
    root.mainloop()